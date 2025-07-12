# classes/views.py
from rest_framework import generics, views, status, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError, NotFound, PermissionDenied
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend

from accounts.permissions import IsBGH, IsGiaoVu
from .models import Khoi, LopHoc, LopHoc_MonHoc, LopHoc_HocSinh
from .serializers import KhoiSerializer, LopHocSerializer, LopHocMonHocUpdateSerializer
from students.serializers import HocSinhSerializer
from students.models import HocSinh
from subjects.models import MonHoc
from subjects.serializers import MonHocSerializer
from grading.models import DiemSo

from configurations.models import ThamSo, NienKhoa

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _is_current_nienkhoa(nienkhoa_id):
    """Hàm helper để kiểm tra niên khóa hiện hành."""
    latest_nien_khoa = NienKhoa.objects.order_by('-TenNienKhoa').first()
    return latest_nien_khoa and nienkhoa_id == latest_nien_khoa.id

# Danh sách khối học (dropdown)
class KhoiListView(generics.ListAPIView):
    queryset = Khoi.objects.all()
    serializer_class = KhoiSerializer
    permission_classes = [IsAuthenticated, IsBGH | IsGiaoVu]

# Tạo và lọc lớp học (admin)
class LopHocListCreateView(generics.ListCreateAPIView):
    queryset = LopHoc.objects.select_related('IDKhoi', 'IDNienKhoa', 'IDToHop').prefetch_related('MonHoc').order_by('-IDNienKhoa__TenNienKhoa', 'TenLop')
    serializer_class = LopHocSerializer
    permission_classes = [IsAuthenticated, IsBGH | IsGiaoVu]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['IDNienKhoa', 'IDKhoi', 'IDToHop']
    search_fields = ['TenLop']

# Danh sách lớp học dùng cho dropdown (lọc đơn giản theo Niên Khóa)
class LopHocListView(generics.ListAPIView):
    serializer_class = LopHocSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = LopHoc.objects.all()
        nienkhoa_id = self.request.query_params.get("nienkhoa_id")
        if nienkhoa_id:
            queryset = queryset.filter(IDNienKhoa__id=nienkhoa_id)
        return queryset

# Chi tiết lớp học
class LopHocDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LopHoc.objects.all()
    serializer_class = LopHocSerializer
    permission_classes = [IsAuthenticated, IsBGH | IsGiaoVu]

    def perform_update(self, serializer):
        if not _is_current_nienkhoa(serializer.instance.IDNienKhoa_id):
            raise PermissionDenied("Chỉ được phép sửa lớp học của niên khóa hiện hành.")
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        if not _is_current_nienkhoa(instance.IDNienKhoa_id):
            raise PermissionDenied("Chỉ được phép xóa lớp học của niên khóa hiện hành.")

        if instance.HocSinh.exists():
            raise ValidationError("Không thể xóa lớp học này vì vẫn còn học sinh trong lớp.")
        instance.delete()

# Cập nhật môn học theo lớp học
class LopHocMonHocUpdateView(views.APIView):
    permission_classes = [IsAuthenticated, IsBGH | IsGiaoVu]

    def get_lop_hoc(self, pk):
        try:
            return LopHoc.objects.get(pk=pk)
        except LopHoc.DoesNotExist:
            raise NotFound("Lớp học không tồn tại.") 

    def post(self, request, pk):
        lop_hoc = self.get_lop_hoc(pk)
        if not lop_hoc:
            return Response({"detail": "Lớp học không tồn tại."}, status=status.HTTP_404_NOT_FOUND)

        if not _is_current_nienkhoa(lop_hoc.IDNienKhoa_id):
             raise PermissionDenied("Chỉ được phép cập nhật môn học cho lớp của niên khóa hiện hành.")

        if DiemSo.objects.filter(IDLopHoc=lop_hoc).exists():
            return Response(
                {"detail": "Không thể thay đổi danh sách môn học vì đã có điểm số được nhập cho lớp này."},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = LopHocMonHocUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        monhoc_ids = serializer.validated_data['monhoc_ids']
        nien_khoa_lop_hoc = lop_hoc.IDNienKhoa
        try:
            tham_so = ThamSo.objects.get(IDNienKhoa=nien_khoa_lop_hoc)
            so_mon_hoc_toi_da = tham_so.SoMonHocToiDa
            
            if len(monhoc_ids) > so_mon_hoc_toi_da:
                return Response(
                    {"detail": f"Số lượng môn học đã chọn ({len(monhoc_ids)}) vượt quá giới hạn cho phép ({so_mon_hoc_toi_da}) của niên khóa."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ThamSo.DoesNotExist:
            return Response(
                {"detail": f"Chưa có quy định về số lượng môn học cho niên khóa {nien_khoa_lop_hoc.TenNienKhoa}."},
                status=status.HTTP_400_BAD_REQUEST
            )
        mon_hoc_qs = MonHoc.objects.filter(pk__in=monhoc_ids)
        
        
        nien_khoa_ids_of_selected_subjects = set(mon_hoc_qs.values_list('IDNienKhoa_id', flat=True))
        if len(nien_khoa_ids_of_selected_subjects) > 1 or (len(nien_khoa_ids_of_selected_subjects) == 1 and nien_khoa_lop_hoc.id not in nien_khoa_ids_of_selected_subjects):
             return Response({
                "detail": f"Một hoặc nhiều môn học không thuộc niên khóa '{nien_khoa_lop_hoc.TenNienKhoa}'."},
                status=status.HTTP_400_BAD_REQUEST)

        
        lop_hoc.MonHoc.set(mon_hoc_qs)
        
        return Response({"message": "Cập nhật môn học cho lớp thành công."}, status=status.HTTP_200_OK)

# Danh sách môn học theo lớp (cho giáo viên nhập điểm)
class MonHocTheoLopView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        lop_hoc_id = request.query_params.get('lop_hoc_id')
        if not lop_hoc_id:
            return Response({"detail": "Thiếu ID lớp học"}, status=400)

        mon_hoc_qs = LopHoc_MonHoc.objects.filter(IDLopHoc_id=lop_hoc_id).select_related('IDMonHoc')
        mon_hoc_list = [mh.IDMonHoc for mh in mon_hoc_qs]
        return Response(MonHocSerializer(mon_hoc_list, many=True).data)


class LopHocHocSinhManagementView(APIView):
    """
    GET: Lấy danh sách học sinh trong lớp và danh sách học sinh có thể được thêm vào lớp.
    POST: Cập nhật danh sách học sinh cho lớp.
    """
    permission_classes = [IsAuthenticated, IsGiaoVu | IsBGH]

    def get_lop_hoc(self, pk):
        try:
            return LopHoc.objects.select_related('IDNienKhoa', 'IDKhoi').get(pk=pk)
        except LopHoc.DoesNotExist:
            raise NotFound(detail="Lớp học không tồn tại.")

    def get(self, request, pk, format=None):
        lop_hoc = self.get_lop_hoc(pk)
        nien_khoa = lop_hoc.IDNienKhoa
        khoi = lop_hoc.IDKhoi

        # Lấy sĩ số tối đa từ quy định
        try:
            tham_so = ThamSo.objects.get(IDNienKhoa=nien_khoa)
            siso_toida = tham_so.SiSoToiDa
        except ThamSo.DoesNotExist:
            return Response(
                {"detail": f"Chưa có quy định về sĩ số cho niên khóa {nien_khoa.TenNienKhoa}."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 1. Lấy danh sách học sinh đã có trong lớp
        students_in_class = lop_hoc.HocSinh.all().order_by('Ten', 'Ho')

        # 2. Lấy danh sách học sinh có thể thêm vào lớp
        # Điều kiện:
        # - Cùng niên khóa tiếp nhận.
        # - Cùng khối dự kiến.
        # - Chưa được xếp vào bất kỳ lớp nào trong niên khóa này.

        # Lấy ID của tất cả học sinh đã được phân lớp trong niên khóa này
        assigned_student_ids = LopHoc_HocSinh.objects.filter(
            IDLopHoc__IDNienKhoa=nien_khoa
        ).values_list('IDHocSinh_id', flat=True)

        students_available = HocSinh.objects.filter(
            IDNienKhoaTiepNhan=nien_khoa,
            KhoiDuKien=khoi
        ).exclude(
            id__in=list(assigned_student_ids)
        ).order_by('Ten', 'Ho')

        # Serialize dữ liệu
        students_in_class_serializer = HocSinhSerializer(students_in_class, many=True)
        students_available_serializer = HocSinhSerializer(students_available, many=True)

        return Response({
            'lop_hoc_info': LopHocSerializer(lop_hoc).data,
            'students_in_class': students_in_class_serializer.data,
            'students_available': students_available_serializer.data,
            'siso_toida': siso_toida,
        })

    def post(self, request, pk, format=None):
        lop_hoc = self.get_lop_hoc(pk)
        student_ids = request.data.get('student_ids', [])

        if not isinstance(student_ids, list):
            return Response({"detail": "Dữ liệu student_ids phải là một danh sách."}, status=status.HTTP_400_BAD_REQUEST)

        # Kiểm tra sĩ số tối đa
        try:
            siso_toida = ThamSo.objects.get(IDNienKhoa=lop_hoc.IDNienKhoa).SiSoToiDa
            if len(student_ids) > siso_toida:
                return Response(
                    {"detail": f"Số lượng học sinh ({len(student_ids)}) vượt quá sĩ số tối đa cho phép ({siso_toida})."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ThamSo.DoesNotExist:
             return Response(
                {"detail": f"Chưa có quy định về sĩ số cho niên khóa {lop_hoc.IDNienKhoa.TenNienKhoa}."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Dùng `set` để cập nhật, DRF sẽ tự động xử lý thêm/xóa
        # Signal `m2m_changed` sẽ được kích hoạt để cập nhật sĩ số
        lop_hoc.HocSinh.set(student_ids)

        return Response({"message": f"Cập nhật danh sách học sinh cho lớp {lop_hoc.TenLop} thành công."}, status=status.HTTP_200_OK)
    

class XuatDanhSachHocSinhView(APIView):
    """
    API để xuất danh sách học sinh của một lớp ra file Excel.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        lophoc_id = request.query_params.get('lophoc_id')
        if not lophoc_id:
            return Response({"detail": "Vui lòng cung cấp ID của lớp học."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            lop_hoc = LopHoc.objects.select_related('IDNienKhoa').get(pk=lophoc_id)
        except LopHoc.DoesNotExist:
            return Response({"detail": "Lớp học không tồn tại."}, status=status.HTTP_404_NOT_FOUND)

        # Lấy sĩ số tối đa để hiển thị
        try:
            tham_so = ThamSo.objects.get(IDNienKhoa=lop_hoc.IDNienKhoa)
            siso_toida = tham_so.SiSoToiDa
        except ThamSo.DoesNotExist:
            siso_toida = "N/A" # Giá trị mặc định nếu không có quy định

        hoc_sinh_list = lop_hoc.HocSinh.all().order_by('Ten', 'Ho')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"DS Lop {lop_hoc.TenLop}"

        title_font = Font(name='Calibri', size=16, bold=True)
        header_font = Font(name='Calibri', size=12, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "DANH SÁCH HỌC SINH"
        title_cell.font = title_font
        title_cell.alignment = center_alignment

        ws.append([]) 
        ws.append(['Niên khóa:', lop_hoc.IDNienKhoa.TenNienKhoa])
        ws['A3'].font = header_font
        ws.append(['Lớp:', lop_hoc.TenLop])
        ws['A4'].font = header_font
        
        # === BỔ SUNG DÒNG SĨ SỐ ===
        ws.append(['Sĩ số:', f"{lop_hoc.SiSo} / {siso_toida}"])
        ws['A5'].font = header_font
        ws.append([]) 

        table_headers = ['STT', 'Họ và tên', 'Giới tính', 'Ngày sinh', 'Email', 'Địa chỉ']
        ws.append(table_headers)
        header_row_num = ws.max_row
        
        for cell in ws[header_row_num]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        for index, hs in enumerate(hoc_sinh_list, start=1):
            row_data = [
                index, f"{hs.Ho} {hs.Ten}", hs.GioiTinh,
                hs.NgaySinh.strftime('%d/%m/%Y'), hs.Email or '', hs.DiaChi
            ]
            ws.append(row_data)
            # Áp dụng border và căn lề cho dòng dữ liệu
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                if cell.column == 1: # Cột STT
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
        
        for col_index in range(1, len(table_headers) + 1):
            column_letter = get_column_letter(col_index)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Danh_sach_lop_{lop_hoc.TenLop}_{lop_hoc.IDNienKhoa.TenNienKhoa}.xlsx"
        response['Content-Disposition'] = f'attachment; filename*="UTF-8\'\'{filename}"' 
        wb.save(response)
        return response
    

class DanhSachHocSinhJsonView(generics.ListAPIView):
    """
    API trả về danh sách học sinh của một lớp dưới dạng JSON để hiển thị trên web.
    """
    serializer_class = HocSinhSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        lophoc_id = self.request.query_params.get('lophoc_id')
        if not lophoc_id:
            return HocSinh.objects.none() # Trả về rỗng nếu không có ID lớp

        try:
            lop_hoc = LopHoc.objects.get(pk=lophoc_id)
            # Lấy danh sách học sinh và sắp xếp theo Tên, Họ
            return lop_hoc.HocSinh.all().order_by('Ten', 'Ho')
        except LopHoc.DoesNotExist:
            return HocSinh.objects.none()