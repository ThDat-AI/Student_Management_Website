from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework import status, generics
from rest_framework.decorators import api_view, permission_classes

from django.http import HttpResponse
import openpyxl
import io
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.cell.cell import Cell

from grading.models import DiemSo, HocKy
from grading.serializers import DiemSoSerializer, HocKySerializer, HocSinhDiemSerializer
from students.models import HocSinh
from classes.models import LopHoc_HocSinh, LopHoc_MonHoc
from configurations.models import ThamSo

from classes.models import LopHoc
from configurations.models import  NienKhoa


# ====== HÀM: Tự động giãn cột Excel ======
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = None
        for cell in col:
            if isinstance(cell, Cell): 
                if column_letter is None:
                    column_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = max_length + 2


# ====== API LẤY DANH SÁCH HỌC SINH VÀ ĐIỂM ======
class DiemSoListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Lấy IDNienKhoa từ query params, nó có thể có hoặc không
        IDNienKhoa = request.query_params.get('IDNienKhoa')
        IDLopHoc = request.query_params.get('IDLopHoc')
        IDMonHoc = request.query_params.get('IDMonHoc')
        IDHocKy = request.query_params.get('IDHocKy')

        if not all([IDLopHoc, IDMonHoc, IDHocKy]):
            return Response({"detail": "Thiếu tham số Lớp học, Môn học hoặc Học kỳ."}, status=400)
        
        # Nếu không có IDNienKhoa được cung cấp, tự động lấy niên khóa mới nhất
        if not IDNienKhoa:
            latest_nien_khoa = NienKhoa.objects.order_by('-TenNienKhoa').first()
            if not latest_nien_khoa:
                return Response({"detail": "Hệ thống chưa có niên khóa nào."}, status=400)
            IDNienKhoa = latest_nien_khoa.id
        
        # Logic kiểm tra môn học thuộc lớp
        if not LopHoc_MonHoc.objects.filter(IDLopHoc_id=IDLopHoc, IDMonHoc_id=IDMonHoc).exists():
            return Response({"detail": "Môn học không thuộc lớp học này"}, status=400)

        # Lọc học sinh theo IDLopHoc và IDNienKhoa đã xác định
        hoc_sinh_ids = LopHoc_HocSinh.objects.filter(IDLopHoc_id=IDLopHoc).values_list("IDHocSinh_id", flat=True)
        hoc_sinh_list = HocSinh.objects.filter(id__in=hoc_sinh_ids, IDNienKhoaTiepNhan_id=IDNienKhoa)

        serializer = HocSinhDiemSerializer(
            hoc_sinh_list,
            many=True,
            context={
                'IDLopHoc': IDLopHoc,
                'IDMonHoc': IDMonHoc,
                'IDHocKy': IDHocKy,
            }
        )
        return Response(serializer.data)


# ====== API CẬP NHẬT HOẶC TẠO ĐIỂM MỚI ======
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cap_nhat_diem(request):
    data = request.data
    required_fields = ['IDHocSinh', 'IDLopHoc', 'IDMonHoc', 'IDHocKy']
    for f in required_fields:
        if f not in data:
            return Response({f: 'Trường này bắt buộc'}, status=400)

    # === LOGIC KIỂM TRA QUYỀN SỬA ĐIỂM ===
    try:
        lop_hoc = LopHoc.objects.select_related('IDNienKhoa__thamso').get(pk=data['IDLopHoc'])
        thamso = lop_hoc.IDNienKhoa.thamso
        id_hoc_ky = int(data['IDHocKy'])

        # Giả định ID Học kỳ 1 là 1, Học kỳ 2 là 2
        if id_hoc_ky == 1 and not thamso.ChoPhepSuaDiemHK1:
            raise PermissionDenied("Hệ thống đã khóa chức năng nhập/sửa điểm cho Học kỳ 1.")
        elif id_hoc_ky == 2 and not thamso.ChoPhepSuaDiemHK2:
            raise PermissionDenied("Hệ thống đã khóa chức năng nhập/sửa điểm cho Học kỳ 2.")

    except (LopHoc.DoesNotExist, ThamSo.DoesNotExist):
        # Nếu không tìm thấy tham số, cho phép sửa để tránh chặn oan
        pass 
    except PermissionDenied as e:
        # Bắt lỗi và trả về cho client
        return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)
    

    diem15 = data.get('Diem15')
    diem1tiet = data.get('Diem1Tiet')

    if not LopHoc_MonHoc.objects.filter(IDLopHoc_id=data['IDLopHoc'], IDMonHoc_id=data['IDMonHoc']).exists():
        return Response({"detail": "Môn học không thuộc lớp học này"}, status=400)

    obj, _ = DiemSo.objects.get_or_create(
        IDHocSinh_id=data['IDHocSinh'],
        IDLopHoc_id=data['IDLopHoc'],
        IDMonHoc_id=data['IDMonHoc'],
        IDHocKy_id=data['IDHocKy']
    )

    obj.Diem15 = diem15
    obj.Diem1Tiet = diem1tiet

    
    if diem15 is not None and diem1tiet is not None:
        try:
            obj.DiemTB = round((float(diem15) + 2 * float(diem1tiet)) / 3, 2)
        except (ValueError, TypeError):
            obj.DiemTB = None
    else:
        obj.DiemTB = None

    obj.save()
    return Response(DiemSoSerializer(obj).data, status=status.HTTP_200_OK)



class ListHocKyView(generics.ListAPIView):
    queryset = HocKy.objects.all()
    serializer_class = HocKySerializer
    permission_classes = [IsAuthenticated]



class XuatExcelDiemSoAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        IDNienKhoa = request.query_params.get("IDNienKhoa")
        IDLopHoc = request.query_params.get("IDLopHoc")
        IDMonHoc = request.query_params.get("IDMonHoc")
        IDHocKy = request.query_params.get("IDHocKy")

        qs = DiemSo.objects.filter(
            IDLopHoc=IDLopHoc,
            IDMonHoc=IDMonHoc,
            IDHocKy=IDHocKy,
            IDHocSinh__IDNienKhoaTiepNhan=IDNienKhoa
        ).select_related("IDHocSinh", "IDLopHoc", "IDMonHoc", "IDHocKy")

        if not qs.exists():
            return HttpResponse("Không có dữ liệu", status=400)

        # Thông tin lọc
        lop_hoc = qs[0].IDLopHoc.TenLop
        mon_hoc = qs[0].IDMonHoc.TenMonHoc
        hoc_ky = qs[0].IDHocKy.TenHocKy
        nien_khoa = qs[0].IDHocSinh.IDNienKhoaTiepNhan.TenNienKhoa

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bảng điểm"

        # Styles
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        fill_khongdat = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Tiêu đề + bộ lọc
        ws.merge_cells("A1:E1")
        ws["A1"] = "BẢNG ĐIỂM MÔN HỌC"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = center_align

        ws.append([])
        ws.append(["Niên khóa:", nien_khoa])
        ws.append(["Lớp học:", lop_hoc])
        ws.append(["Môn học:", mon_hoc])
        ws.append(["Học kỳ:", hoc_ky])
        ws.append([])

        # Header
        headers = ["Họ tên", "Điểm 15 phút", "Điểm 1 tiết", "Điểm TB", "Kết quả"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_align

        # Điểm đạt môn
        diem_dat_mon = ThamSo.objects.last().DiemDatMon if ThamSo.objects.exists() else 5.0

        # Dữ liệu học sinh
        for d in qs:
            tb = (d.Diem15 + 2 * d.Diem1Tiet) / 3 if d.Diem15 is not None and d.Diem1Tiet is not None else None
            ket_qua = "Đạt" if tb is not None and tb >= diem_dat_mon else "Không đạt"
            ho_ten = f"{d.IDHocSinh.Ho} {d.IDHocSinh.Ten}"

            row_data = [
                ho_ten,
                d.Diem15 or "",
                d.Diem1Tiet or "",
                f"{tb:.2f}" if tb is not None else "",
                ket_qua
            ]
            ws.append(row_data)

            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = center_align
                if ket_qua == "Không đạt":
                    cell.fill = fill_khongdat

        auto_adjust_column_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bang_diem.xlsx"},
        )
