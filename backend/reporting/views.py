from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import io
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.cell.cell import Cell
from django.core.exceptions import ObjectDoesNotExist

from grading.models import DiemSo, HocKy
from classes.models import LopHoc, LopHoc_MonHoc
from configurations.models import ThamSo
from subjects.models import MonHoc


# ======= STYLE CHUNG =========
def style_excel_header(ws, headers):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    font = Font(bold=True)
    align = Alignment(horizontal='center', vertical='center')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = font
        cell.border = border
        cell.alignment = align


def style_excel_rows(ws, data_start_row=6):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            cell.border = border
            cell.alignment = align


def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = None
        for cell in col:
            if isinstance(cell, Cell):
                if column_letter is None:
                    column_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = max_length + 2


# ======== API BÁO CÁO MÔN HỌC =========
class BaoCaoMonHocView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        IDMonHoc = request.query_params.get("IDMonHoc")
        IDHocKy = request.query_params.get("IDHocKy")

        if not IDMonHoc or not IDHocKy:
            return Response({"detail": "Thiếu IDMonHoc hoặc IDHocKy"}, status=400)

        lop_ids = LopHoc_MonHoc.objects.filter(IDMonHoc=IDMonHoc).values_list("IDLopHoc", flat=True).distinct()
        lop_list = LopHoc.objects.filter(id__in=lop_ids)

        data = []
        for lop in lop_list:
            diem_list = DiemSo.objects.filter(IDLopHoc=lop.id, IDMonHoc=IDMonHoc, IDHocKy=IDHocKy)
            si_so = diem_list.count()
            so_luong_dat = 0

            for diem in diem_list:
                dtb = (diem.Diem15 + 2 * diem.Diem1Tiet) / 3 if diem.Diem15 is not None and diem.Diem1Tiet is not None else None
                try:
                    diem_dat_mon = ThamSo.objects.get(IDNienKhoa=diem.IDHocSinh.IDNienKhoaTiepNhan).DiemDatMon
                except ThamSo.DoesNotExist:
                    diem_dat_mon = 5.0
                if dtb is not None and dtb >= diem_dat_mon:
                    so_luong_dat += 1

            ti_le = (so_luong_dat / si_so) * 100 if si_so > 0 else 0
            if si_so > 0:
                data.append({
                    "TenLop": lop.TenLop,
                    "SiSo": si_so,
                    "SoLuongDat": so_luong_dat,
                    "TiLe": round(ti_le, 2),
                })

        return Response(data)


class ExportBaoCaoMonHocExcel(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        IDMonHoc = request.query_params.get("IDMonHoc")
        IDHocKy = request.query_params.get("IDHocKy")
        IDNienKhoa = request.query_params.get("IDNienKhoa")

        request._request.path = "/api/reporting/baocao/monhoc/"
        response = BaoCaoMonHocView().get(request)
        data = response.data

        # === Đảm bảo lấy đúng tên môn học ===
        try:
            ten_mon = MonHoc.objects.get(id=int(IDMonHoc)).TenMonHoc
        except (ValueError, ObjectDoesNotExist):
            ten_mon = "Không rõ"

        try:
            ten_hocky = HocKy.objects.get(id=int(IDHocKy)).TenHocKy
        except:
            ten_hocky = f"Học kỳ {IDHocKy}"

        try:
            thamso = ThamSo.objects.get(IDNienKhoa=int(IDNienKhoa))
            nien_khoa = thamso.IDNienKhoa.TenNienKhoa
        except:
            nien_khoa = "Không rõ"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo môn học"

        ws.merge_cells("A1:D1")
        ws["A1"] = "BÁO CÁO TỔNG KẾT MÔN HỌC"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Niên khóa: {nien_khoa}"
        ws["B2"] = f"Môn học: {ten_mon}"
        ws["C2"] = f"Học kỳ: {ten_hocky}"
        for col in ["A2", "B2", "C2"]:
            ws[col].font = Font(italic=True)

        headers = ["Lớp", "Sĩ số", "Số lượng đạt", "Tỉ lệ (%)"]
        style_excel_header(ws, headers)

        for idx, row in enumerate(data, start=6):
            ws.append([row["TenLop"], row["SiSo"], row["SoLuongDat"], row["TiLe"]])

        style_excel_rows(ws, data_start_row=6)
        auto_adjust_column_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bao_cao_mon_hoc.xlsx"},
        )


# ========= API BÁO CÁO HỌC KỲ ==========
class BaoCaoHocKyView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        IDNienKhoa = request.query_params.get("IDNienKhoa")
        IDHocKy = request.query_params.get("IDHocKy")

        if not IDNienKhoa or not IDHocKy:
            return Response({"detail": "Thiếu IDNienKhoa hoặc IDHocKy"}, status=400)

        lop_list = LopHoc.objects.filter(IDNienKhoa=IDNienKhoa)
        data = []

        for lop in lop_list:
            diem_list = DiemSo.objects.filter(IDLopHoc=lop.id, IDHocKy=IDHocKy)
            hoc_sinh_ids = diem_list.values_list("IDHocSinh", flat=True).distinct()
            si_so = len(hoc_sinh_ids)
            so_luong_dat = 0

            for hs_id in hoc_sinh_ids:
                diem_hs = diem_list.filter(IDHocSinh_id=hs_id)
                diem_tbs = [(d.Diem15 + 2 * d.Diem1Tiet) / 3 for d in diem_hs if d.Diem15 is not None and d.Diem1Tiet is not None]
                if not diem_tbs:
                    continue

                diem_tb_hocky = sum(diem_tbs) / len(diem_tbs)

                try:
                    diem_dat_mon = ThamSo.objects.get(IDNienKhoa=IDNienKhoa).DiemDatMon
                except ThamSo.DoesNotExist:
                    diem_dat_mon = 5.0

                if diem_tb_hocky >= diem_dat_mon:
                    so_luong_dat += 1

            ti_le = (so_luong_dat / si_so) * 100 if si_so > 0 else 0
            if si_so > 0:
                data.append({
                    "TenLop": lop.TenLop,
                    "SiSo": si_so,
                    "SoLuongDat": so_luong_dat,
                    "TiLe": round(ti_le, 2),
                })

        return Response(data)


class ExportBaoCaoHocKyExcel(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        IDHocKy = request.query_params.get("IDHocKy")
        IDNienKhoa = request.query_params.get("IDNienKhoa")

        request._request.path = "/api/reporting/baocao/hocky/"
        response = BaoCaoHocKyView().get(request)
        data = response.data

        try:
            ten_hocky = HocKy.objects.get(id=int(IDHocKy)).TenHocKy
        except:
            ten_hocky = f"Học kỳ {IDHocKy}"

        try:
            thamso = ThamSo.objects.get(IDNienKhoa=int(IDNienKhoa))
            nien_khoa = thamso.IDNienKhoa.TenNienKhoa
        except:
            nien_khoa = "Không rõ"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo học kỳ"

        ws.merge_cells("A1:D1")
        ws["A1"] = "BÁO CÁO TỔNG KẾT HỌC KỲ"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Niên khóa: {nien_khoa}"
        ws["B2"] = f"Học kỳ: {ten_hocky}"
        for col in ["A2", "B2"]:
            ws[col].font = Font(italic=True)

        headers = ["Lớp", "Sĩ số", "Số lượng đạt", "Tỉ lệ (%)"]
        style_excel_header(ws, headers)

        for idx, row in enumerate(data, start=6):
            ws.append([row["TenLop"], row["SiSo"], row["SoLuongDat"], row["TiLe"]])

        style_excel_rows(ws, data_start_row=6)
        auto_adjust_column_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bao_cao_hoc_ky.xlsx"},
        )
